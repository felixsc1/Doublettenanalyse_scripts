import openpyxl
import pandas as pd
import os

from helper_functions.hardcoded_values import produkte_dict_personen, servicerollen, produkte_dict
from .file_io_functions import load_data
import pickle
import glob
import numpy as np
import ast
import re


def extract_hyperlinks(file_path, columns):
    """
    Extract hyperlinks from specified columns in an Excel file and add new columns with the suffix `_link`.

    Parameters:
        file_path (str): Path to the Excel file.
        columns (list): List of column names to extract hyperlinks from.

    Returns:
        None. The function saves the updated DataFrame to a new file with a `_hyperlinks` suffix.
    """
    # Load the Excel file using openpyxl
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Dictionary to store hyperlinks for each column
    hyperlink_dict = {}
    for column_name in columns:
        col_idx = None
        for idx, col in enumerate(sheet.iter_cols(1, sheet.max_column)):
            if col[0].value == column_name:
                col_idx = idx + 1  # 1-based index
                break

        if col_idx is None:
            print(f"Column '{column_name}' not found.")
            continue

        # Extract hyperlinks
        hyperlinks = []
        for row in sheet.iter_rows(
            min_row=2, max_row=sheet.max_row, min_col=col_idx, max_col=col_idx
        ):
            cell = row[0]
            if cell.hyperlink:
                hyperlinks.append(cell.hyperlink.target)
            else:
                hyperlinks.append(None)

        hyperlink_dict[column_name + "_link"] = hyperlinks

    # Load the Excel data into pandas DataFrame, then add the hyperlinks
    df = pd.read_excel(file_path, engine="openpyxl")
    for col_name, links in hyperlink_dict.items():
        df[col_name] = links

    # Determine the save path
    base_name = os.path.basename(file_path)
    name_without_extension = os.path.splitext(base_name)[0]
    save_name = name_without_extension + "_hyperlinks.xlsx"
    save_path = os.path.join(os.path.dirname(file_path), save_name)

    # Save the updated DataFrame back to Excel with the new name
    df.to_excel(save_path, index=False, engine="openpyxl")

    return save_path


def normalize_string(string_in):
    # Normalize Names and Addresses: lowercase, strip whitespace, replace multiple whitespace with single whitespace
    normalized = string_in.lower().strip()
    normalized = re.sub(r"\s+", " ", normalized)
    return normalized


def basic_cleanup(df, remove_personen_Sonstiges=True):
    """
    Performs some basic corrections to String formatting.
    Removes Inaktiv entries and Personen with Sonstiges Verknüpfungsart.
    """

    # Remove rows missing a name.
    df_cleaned = df[df["Name"].apply(lambda x: isinstance(x, str))]

    df_cleaned["Name_original"] = df_cleaned["Name"]  # Keep original for reference
    df_cleaned["Name"] = df_cleaned["Name"].apply(normalize_string)

    # sort by Name
    df_cleaned = df_cleaned.sort_values("Name")

    df_cleaned = df_cleaned.replace(
        {pd.NA: "", "nan": ""}
    )  # sometimes cells contain string 'nan' this causes problems later

    # BUG: Remember that .astype(str) will replace pd.NA with "nan".

    # removing all spaces between numbers, because they are placed quite inconsistently.
    df_cleaned["Telefonnummer"] = (
        df_cleaned["Telefonnummer"].str.replace(" ", "").astype(str)
    )
    # email addresses will also need some processing. for now, ensure values are strings and lowercase.
    df_cleaned["EMailAdresse"] = df_cleaned["EMailAdresse"].astype(str).str.lower()

    # Ensure that PLZ is string, for some reason its sometimes float which causes problems
    df_cleaned["ZipPostalCode"] = df_cleaned["ZipPostalCode"].astype(str)

    # Filter out all Inaktive and Sonstiges Verknüpfungsart for Personen
    df_cleaned = df_cleaned[df_cleaned["Aktiv"] != False]
    if remove_personen_Sonstiges:
        df_cleaned = df_cleaned[df_cleaned["Verknuepfungsart"] != "Sonstiges"]

    return df_cleaned


def construct_address_string(row, organisation=False):
    """
    expects row to have the elements listed below.
    Since extra text in address1 and address2 can confuse gmaps, also return partial address with only street and number (currently not used anymore).
    """
    # Check if ZipPostalCode is a number
    # Check if ZipPostalCode is not NaN and not the string 'nan'
    zip_code = row["ZipPostalCode"]
    if pd.notna(zip_code) and str(zip_code).lower() != "nan" and zip_code != "":
        try:
            zip_postal_code = str(int(float(zip_code)))
        except ValueError:
            zip_postal_code = str(zip_code)  # if it has letters, e.g. UK
    elif organisation:
        korr_zip_code = row["Korr_ZipPostalCode"]
        if (
            pd.notna(korr_zip_code)
            and str(korr_zip_code).lower() != "nan"
            and korr_zip_code != ""
        ):
            try:
                zip_postal_code = str(int(float(korr_zip_code)))
            except ValueError:
                zip_postal_code = str(korr_zip_code)
        else:
            zip_postal_code = ""
    else:
        zip_postal_code = ""

    address_elements = [
        str(row["Street"]),
        str(row["HouseNumber"]),
        str(row["Address1"]),
        str(row["Address2"]),
        str(row["PostOfficeBox"]),
        zip_postal_code,
        str(row["City"]),
        str(row["CountryName"]),
    ]

    address_elements_partial = [
        str(row["Street"]),
        str(row["HouseNumber"]),
        zip_postal_code,
        str(row["City"]),
        str(row["CountryName"]),
    ]

    elements_without_zip_code = [
        str(row["Street"]),
        str(row["HouseNumber"]),
        str(row["Address1"]),
        str(row["Address2"]),
        str(row["PostOfficeBox"]),
        str(row["City"]),
        str(row["CountryName"]),
    ]

    # Check if all address elements are NaN (or "nan" or empty strings), try use korrespondenz_adresse instead
    if all(
        pd.isna(element) or element == "" or element.lower() == "nan"
        for element in elements_without_zip_code
    ):
        if organisation:  # Personen don't have these columns
            address_elements = [
                str(row["Korr_Street"]),
                str(row["Korr_HouseNumber"]),
                str(row["Korr_Address1"]),
                str(row["Korr_Address2"]),
                str(row["Korr_PostOfficeBox"]),
                zip_postal_code,
                str(row["Korr_City"]),
                str(row["Korr_CountryName"]),
            ]

            address_elements_partial = [
                str(row["Korr_Street"]),
                str(row["Korr_HouseNumber"]),
                zip_postal_code,
                str(row["Korr_City"]),
                str(row["Korr_CountryName"]),
            ]
        else:
            return pd.Series(["", ""])
        # it that is also empty, return empty string
        if all(
            pd.isna(element) or element == "" or element.lower() == "nan"
            for element in address_elements
        ):
            return pd.Series(["", ""])

    # Filter out None, 'nan', and empty strings, then join with commas
    full_address = ", ".join(
        filter(lambda x: x and x != "nan" and str(x).strip(), address_elements)
    )

    partial_address = ", ".join(
        filter(lambda x: x and x != "nan" and str(x).strip(), address_elements_partial)
    )

    # Finally make it lowercase, remove additional spaces
    full_address = normalize_string(full_address)
    partial_address = normalize_string(partial_address)

    output_columns = [full_address, partial_address]

    return pd.Series(output_columns)


def aggregate_identical_UIDs(df):
    """
    Those with identical IDs that are a result of flattening from linq output.
    Does not affect duplicated with same name and different IDs.
    VerknüpftesObjekt etc. are aggregated as lists.
    """

    def aggregate_to_list(series):
        return series.tolist()

    def first_entry(series):
        return series.iloc[0]

    columns_as_lists = [
        "Verknuepfungsart",
        "VerknuepftesObjektID",
        "VerknuepftesObjekt",
    ]

    aggregation = {
        col: aggregate_to_list if col in columns_as_lists else first_entry
        for col in df.columns
        if col != "ReferenceID"
    }

    grouped = df.groupby("ReferenceID").agg(aggregation).reset_index()

    return grouped


def add_servicerole_column_string(df_data, serviceroles_df):
    # This one adds the actual name of the service role. For personen to give Ausweis higher score later on.
    df_data["Servicerole_string"] = ""

    for index, row in df_data.iterrows():
        ref_id = row["ReferenceID"]
        matching_roles = serviceroles_df[
            serviceroles_df["Rechtsträger_RefID"] == ref_id
        ]

        roles = []
        for _, role_row in matching_roles.iterrows():
            role_ref_id = role_row["ServiceRoleReferenceID"]
            role = servicerollen.get(role_ref_id, "")
            if role:
                roles.append(role)

        df_data.at[index, "Servicerole_string"] = ", ".join(roles)

    return df_data


def add_servicerole_column(df_organisationen, serviceroles_df):
    # Very similar to "add_Produkte_columns()"
    # This one just adds a count, to filter out those that have a service role.
    df_organisationen["Servicerole_count"] = 0

    for ref_id in df_organisationen["ReferenceID"]:
        role_count = serviceroles_df["Rechtsträger_RefID"].eq(ref_id).sum()
        df_organisationen.loc[
            df_organisationen["ReferenceID"] == ref_id, "Servicerole_count"
        ] += role_count
    return df_organisationen


def add_produkte_columns(df_organisationen, organisationsrollen_df):
    # Calculate counts for 'Produkt_Inhaber'
    inhaber_counts = organisationsrollen_df["Inhaber_RefID"].value_counts()
    # Map these counts to df_organisationen
    df_organisationen["Produkt_Inhaber"] = (
        df_organisationen["ReferenceID"].map(inhaber_counts).fillna(0).astype(int)
    )

    # Calculate counts for 'Produkt_Adressant'

    # Calculate counts for each reference ID separately, filling missing values with 0
    rechnungsempfaenger_counts = organisationsrollen_df["Rechnungsempfaenger_RefID"].value_counts()
    korrespondenzempfaenger_counts = organisationsrollen_df["Korrespondenzempfaenger_RefID"].value_counts()
    # Combine the Series, ensuring that missing entries in one Series do not nullify existing counts from the other
    adressant_counts = rechnungsempfaenger_counts.add(korrespondenzempfaenger_counts, fill_value=0)
    # Map these combined counts to df_organisationen
    df_organisationen["Produkt_Adressant"] = df_organisationen["ReferenceID"].map(adressant_counts).fillna(0).astype(int)

    return df_organisationen


def add_personen_produkte_columns(df_data, df_produktrollen):
    # Creating dictionaries for quick lookup with lists to handle non-unique indices
    kontaktperson_dict = (
        df_produktrollen.groupby("Kontaktperson_RefID")
        .apply(lambda x: x.to_dict("records"))
        .to_dict()
    )
    technikperson_dict = (
        df_produktrollen.groupby("Technikperson_RefID")
        .apply(lambda x: x.to_dict("records"))
        .to_dict()
    )
    statistikperson_dict = (
        df_produktrollen.groupby("Statistikperson_RefID")
        .apply(lambda x: x.to_dict("records"))
        .to_dict()
    )

    # Initialize new columns as lists
    df_data["Produkt_rolle"] = [[] for _ in range(len(df_data))]
    df_data["Produkt_RefID"] = [[] for _ in range(len(df_data))]

    # Iterate over each row in df_data
    for index, row in df_data.iterrows():
        reference_id = row["ReferenceID"]

        # Check for matches in dictionaries
        for ref_dict, role in [
            (kontaktperson_dict, "Kontaktperson"),
            (technikperson_dict, "Technikperson"),
            (statistikperson_dict, "Statistikperson"),
        ]:
            if reference_id in ref_dict:
                for produktrollen_row in ref_dict[reference_id]:
                    # Lookup in dictionary and append role with additional information
                    full_id = produktrollen_row["FullID"]
                    additional_info = produkte_dict_personen.get(full_id, "")
                    role_with_info = (
                        f"{role} ({additional_info})" if additional_info else role
                    )

                    df_data.at[index, "Produkt_rolle"].append(role_with_info)
                    df_data.at[index, "Produkt_RefID"].append(
                        produktrollen_row["Produkt_RefID"]
                    )

    return df_data


def add_organisationen_produkte_columns(df_data, df_produktrollen):
    """
    Can be run after add_personen_produkte_columns()
    Is actually meant for Personen. 
    This adds in addition to just Statistikperson etc. also Inhaber, Rechnungsempfänger etc. in the same list in the same columns.
    """
    # Creating dictionaries for quick lookup with lists to handle non-unique indices
    inhaber_dict = (
        df_produktrollen.groupby("Inhaber_RefID")
        .apply(lambda x: x.to_dict("records"))
        .to_dict()
    )
    rechnungsempfaenger_dict = (
        df_produktrollen.groupby("Rechnungsempfaenger_RefID")
        .apply(lambda x: x.to_dict("records"))
        .to_dict()
    )
    korrespondenzempfaenger_dict = (
        df_produktrollen.groupby("Korrespondenzempfaenger_RefID")
        .apply(lambda x: x.to_dict("records"))
        .to_dict()
    )
    
    # Debug: Check if dictionaries are populated
    # print(f"Inhaber Dict: {len(inhaber_dict)} entries")
    # print(f"Rechnungsempfaenger Dict: {len(rechnungsempfaenger_dict)} entries")
    # print(f"Korrespondenzempfaenger Dict: {len(korrespondenzempfaenger_dict)} entries")

    # Ensure columns exist
    if "Produkt_rolle" not in df_data.columns:
        df_data["Produkt_rolle"] = [[] for _ in range(len(df_data))]
    if "Produkt_RefID" not in df_data.columns:
        df_data["Produkt_RefID"] = [[] for _ in range(len(df_data))]

    # Iterate over each row in df_data
    for index, row in df_data.iterrows():
        reference_id = row["ReferenceID"]

        # Check for matches in dictionaries
        for ref_dict, role in [
            (inhaber_dict, "Inhaber"),
            (rechnungsempfaenger_dict, "Rechnungsempfaenger"),
            (korrespondenzempfaenger_dict, "Korrespondenzempfaenger"),
        ]:
            if reference_id in ref_dict:
                for produktrollen_row in ref_dict[reference_id]:
                    # Lookup in dictionary and append role with additional information
                    full_id = produktrollen_row["FullID"]
                    additional_info = produkte_dict.get(full_id, "")
                    role_with_info = (
                        f"{role} ({additional_info})" if additional_info else role
                    )

                    df_data.at[index, "Produkt_rolle"].append(role_with_info)
                    df_data.at[index, "Produkt_RefID"].append(
                        produktrollen_row["Produkt_RefID"]
                    )
                    # print(f"Added {role_with_info} to index {index}")

    return df_data


def get_geschaeftspartner(input_df, folder_path):
    """
    Check if input df has matching ReferenceID with any of the other dfs.
    df gets a new column "Geschaeftspartner" which contains a list of all matching partners.
    """
    # Create the "Geschaeftspartner" column in the input_df
    # print(f"Starting get_geschaeftspartner function with {len(input_df)} rows in input_df")
    input_df["Geschaeftspartner"] = [[] for _ in range(len(input_df))]

    # List all xlsx files in the specified directory
    xlsx_files = glob.glob(f"{folder_path}/*.xlsx")
    # print(f"Found {len(xlsx_files)} xlsx files in {folder_path}")

    # Helper function to check if a ReferenceID exists in any of the dfs and return its name(s)
    def check_reference(reference, df, partner_name):
        if reference in df["ReferenceID"].values:
            # print(f"Found {partner_name} for {reference}")
            return [partner_name]
        return []

    # Load each xlsx file and check for a match with the ReferenceID in input_df
    for xlsx_file in xlsx_files:
        # Extract the partner name from the file name
        partner_name = (
            os.path.basename(xlsx_file)
            .rsplit("-", 1)[-1]
            .rsplit("_", 1)[-1]
            .split(".")[0]
        )
        # print(f"Processing file: {xlsx_file}, partner_name: {partner_name}")

        # Load the dataframe
        df = pd.read_excel(xlsx_file)
        # print(f"Loaded dataframe with {len(df)} rows")

        # Loop through each row in input_df and populate the "Geschaeftspartner" column
        matches_found = 0
        for index, row in input_df.iterrows():
            partners = check_reference(row["ReferenceID"], df, partner_name)
            if partners:
                matches_found += 1
            input_df.at[index, "Geschaeftspartner"].extend(partners)

        # print(f"Found {matches_found} matches for {partner_name}")

    return input_df


def get_true_lists_generic(df):
    """
    Warnung: Speichern zu excel verwandelt Zellen die eine Liste als value enthalten in Strings: ['a','b'] --> "['a','b']"
    Hier fixt dies indem es jede Kolonne die im obigen string format ist umwandelt und in "kolonnename_list" speichert.
    Muss bei jedem einlesen von xlsx daten neu ausgeführt werden.
    """

    def convert_to_list(s):
        try:
            return ast.literal_eval(s)
        except (ValueError, SyntaxError):
            return s  # return the original value if conversion fails

    for col in df.columns:
        # Skip columns that already have a '_list' suffix
        if col.endswith("_list"):
            continue

        # Check if any element in the column is a string that looks like a list
        if (
            df[col]
            .apply(
                lambda x: isinstance(x, str) and x.startswith("[") and x.endswith("]")
            )
            .any()
        ):
            new_col_name = f"{col}_list"
            df[new_col_name] = df[col].apply(convert_to_list)

    return df


def calculate_scores_organisationen(df):
    # new: requires serviceroles and produkte to be integrated.
    df["Debitornummer"] = df["Debitornummer"].fillna(0)
    df["Versandart"] = df["Versandart"].fillna(0)
    df["AnzahlGeschaeftsobjekte"] = df["AnzahlGeschaeftsobjekte"].fillna(0)
    df["AnzahlObjektZeiger"] = df["AnzahlObjektZeiger"].fillna(0)
    df["Debitornummer_check"] = df["Debitornummer"].apply(lambda x: 1 if x > 0 else 0)
    df["UID_CHID_check"] = df["UID_CHID"].apply(
        lambda x: 1 if isinstance(x, str) else 0
    )

    def calculate_score_and_details(row):
        score_components = {
            "Debitornummer": row["Debitornummer_check"] * 100,
            "UID_CHID": row["UID_CHID_check"] * 200,
            "Versandart": 100 if row["Versandart"] == "Portal" else 0,
            "Geschaeftsobjekte": row["AnzahlGeschaeftsobjekte"] * 30,
            "ObjektZeiger": min(row["AnzahlObjektZeiger"] * 10, 100),
            "Verknuepfungsart": sum(
                100 if val == "Administrator" else 50 if val == "Mitarbeiter" else 0
                for val in row["Verknuepfungsart_list"]
            ),
            "Geschaeftspartner": sum(100 for _ in row["Geschaeftspartner_list"]),
            "Produkt_Inhaber": min(row["Produkt_Inhaber"] * 80, 200),
            "Produkt_Adressant": min(row["Produkt_Adressant"] * 30, 100),
            "Servicerole": row["Servicerole_count"] * 50,
            "UID_MASTER": 1000 if row["UID_MASTER"] == True else 0
        }
        
        total_score = sum(score_components.values())
        score_details = ", ".join([f"{k}: {v}" for k, v in score_components.items() if v > 0])
        
        return total_score, score_details

    df["score"], df["score_details"] = zip(*df.apply(calculate_score_and_details, axis=1))
    
    return df


def calculate_scores_personen(df, physisch=False):
    # For Doubletten physisch, UID is not really important. We still consider it here but divided by 10.

    # Fill missing values for non-list columns
    df.fillna(
        {
            "AnzahlGeschaeftsobjekte": 0,
            "Versandart": 0,
            "AnzahlObjektZeiger": 0,
            "AnzahlVerknuepfungen": 0,
            "Servicerole_string": "",
        },
        inplace=True,
    )

    # UID_CHID_check calculation
    df["UID_CHID_check"] = df["UID_CHID"].apply(
        lambda x: (
            0
            if pd.isna(x) or x == ""
            else 1 if str(x).lower() == "notregisteredchid" else 2
        )
    )

    # Function to calculate score and score details
    def score_and_details(row):
        score_components = {
            "Geschaeftsobjekte": row["AnzahlGeschaeftsobjekte"] * 30,
            "UID": int(row["UID_CHID_check"] * 50 / (10 if physisch else 1)),
            "Verknuepfungsart": sum(
                100 if val == "Administrator" else 50 if val == "Mitarbeiter" else 0
                for val in (row["Verknuepfungsart_list"] or [])
            ),
            "Versandart": 100 if row["Versandart"] == "Portal" else 0,
            "ObjektZeiger": np.minimum(row["AnzahlObjektZeiger"] * 10, 100),
            "Geschaeftspartner": sum(
                100 for _ in (row["Geschaeftspartner_list"] or [])
            ),
            "Servicerole_string": 100 if "Ausweis" in row["Servicerole_string"] else 0,
            "Produktrolle": (
                len(row["Produkt_rolle"]) * 100 if row["Produkt_rolle"] else 0
            ),
        }

        if row["EMailAdresse"] and not pd.isna(row["EMailAdresse"]):
            score_components["Email"] = 20

        if row["Telefonnummer"] and not pd.isna(row["Telefonnummer"]):
            score_components["Email"] = 10

        score_details = ", ".join(
            [f"{name} {score}" for name, score in score_components.items() if score > 0]
        )

        total_score = sum(score_components.values())

        return total_score, score_details

    # Apply the function to each row
    df[["score", "score_details"]] = df.apply(
        lambda row: score_and_details(row), axis=1, result_type="expand"
    )

    return df


def raw_cleanup(file_paths_original, raw_data_directory, remove_personen_Sonstiges=True, skip_hyperlink_step=False):
    """
    Main function that calls basic_cleanup(), and various others.
    Also integrates other Expertensuchen such as Serviceroles into df_personen, df_organisationen.
    """
    
    file_paths = file_paths_original.copy()

    # create files with separate hyperlinks column and use this from here on
    print("Reading excel files and extracting hyperlinks (takes several minutes)...")
    if skip_hyperlink_step:
        base_name = os.path.basename(file_paths["organisationen"])
        name_without_extension = os.path.splitext(base_name)[0]
        save_name = name_without_extension + "_hyperlinks.xlsx"
        file_paths["organisationen"] = os.path.join(
            os.path.dirname(file_paths["organisationen"]), save_name
        )

        base_name = os.path.basename(file_paths["personen"])
        name_without_extension = os.path.splitext(base_name)[0]
        save_name = name_without_extension + "_hyperlinks.xlsx"
        file_paths["personen"] = os.path.join(
            os.path.dirname(file_paths["personen"]), save_name
        )

    else:
        df_organisationen = extract_hyperlinks(
            file_paths["organisationen"], ["Objekt", "VerknuepftesObjekt"]
        )
        df_personen = extract_hyperlinks(
            file_paths["personen"], ["Objekt", "VerknuepftesObjekt"]
        )

        # Update the file_paths dictionary
        file_paths["personen"] = df_personen
        file_paths["organisationen"] = df_organisationen

    df_organisationen = load_data(file_paths["organisationen"])
    df_personen = load_data(file_paths["personen"])

    print("Basic cleanup Organisationen & Personen...")
    df_organisationen = basic_cleanup(
        df_organisationen, remove_personen_Sonstiges=False
    )
    df_personen = basic_cleanup(
        df_personen, remove_personen_Sonstiges=remove_personen_Sonstiges
    )

    print("Aggregating additional Expertensuchen...")
    df_organisationen = aggregate_identical_UIDs(df_organisationen)
    df_personen = aggregate_identical_UIDs(df_personen)

    df_organisationen[["address_full", "address_partial"]] = df_organisationen.apply(
        lambda row: construct_address_string(row, organisation=True), axis=1
    )

    df_personen[["address_full", "address_partial"]] = df_personen.apply(
        lambda row: construct_address_string(row, organisation=False), axis=1
    )
    
    # Debugging:
    # temp = df_personen.apply(
    #     lambda row: construct_address_string(row, organisation=False), axis=1
    # )
    # if True:
    #     return temp

    df_organisationen["Name_Zeile2"] = df_organisationen.apply(
        lambda x: (
            x["Name"] + "|" + str(x["Zeile2"])
            if pd.notna(x["Zeile2"]) and x["Zeile2"] != ""
            else x["Name"]
        ),
        axis=1,
    )

    personenservicerolle_df = load_data(file_paths["personenservicerolle"])
    organisationservicerolle_df = load_data(file_paths["organisationservicerolle"])
    df_personen = add_servicerole_column_string(df_personen, personenservicerolle_df)
    df_organisationen = add_servicerole_column_string(
        df_organisationen, organisationservicerolle_df
    )
    df_organisationen = add_servicerole_column(
        df_organisationen, organisationservicerolle_df
    )  # only for score

    # PRODUKTE / ROLLEN
    organisationsrollen_df_1 = load_data(file_paths["organisationsrollen"])
    organisationsrollenFDA_df = load_data(file_paths["organisationsrollenFDA"])
    organisationsrollen_df = pd.concat(
        [organisationsrollen_df_1, organisationsrollenFDA_df], ignore_index=True
    )
    df_organisationen = add_produkte_columns(
        df_organisationen, organisationsrollen_df
    )  # only the count. only needed for score.

    df_personenrollen = load_data(file_paths["personenrollen"])
    df_personen = add_personen_produkte_columns(
        df_personen, df_personenrollen
    )  # Technikperson, Statistikperson, etc.
    df_personen = add_organisationen_produkte_columns(df_personen, organisationsrollen_df) # Inhaber, Rechnungsempfänger, etc.

    # GESCHÄFTSPARTNER
    df_organisationen = get_geschaeftspartner(
        df_organisationen,
        os.path.join(raw_data_directory, "mandanten/organisationen"),
    )
    df_personen = get_geschaeftspartner(df_personen, os.path.join(raw_data_directory, "mandanten/personen"))

    # Processing list column, to have both string and true list representation.
    columns_to_convert = [
        "VerknuepftesObjektID",
        "VerknuepftesObjekt",
        "Verknuepfungsart",
        "Geschaeftspartner",
    ]
    for col in columns_to_convert:
        df_organisationen[col] = df_organisationen[col].apply(str)
        df_personen[col] = df_personen[col].apply(str)
    df_organisationen = get_true_lists_generic(df_organisationen)
    df_personen = get_true_lists_generic(df_personen)

    # Now we have information to calculate scores
    df_organisationen = calculate_scores_organisationen(df_organisationen)
    df_personen = calculate_scores_personen(df_personen)

    # Store dataframes as pickle
    dfs = {
        "personen": df_personen,
        "organisationen": df_organisationen,
        "organisationsrollen": organisationsrollen_df,
    }

    print("Storing dataframes as pickle...")
    # Create the directory if it doesn't exist
    directory = "data/calculated"
    os.makedirs(directory, exist_ok=True)
    if remove_personen_Sonstiges == False:
        filepath = os.path.join(
            directory,
            "personen_organisationen_dfs_processed_with_sonstiges_personen.pickle",
        )
    else:
        filepath = os.path.join(
            directory, "personen_organisationen_dfs_processed.pickle"
        )

    with open(filepath, "wb") as file:
        pickle.dump(dfs, file)

    return df_organisationen, df_personen
